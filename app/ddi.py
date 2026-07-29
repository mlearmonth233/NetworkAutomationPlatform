"""DDI (DNS/DHCP/IPAM) analysis: extracts VLAN interface / DHCP helper-address
configuration from a core switch's running-config, whether that config was
uploaded directly or collected live over SSH. Mirrors the structure of a
real DHCP-info workbook shared by the user: a flattened VLAN/helper table, a
condensed SVI config view, and a deduplicated helper-address summary.

Honest scope note: the flattened table, condensed config view, and helper
occurrence counts are all derived purely from the switch's own
configuration. A hostname for each helper IP is attempted via a best-effort
reverse DNS lookup. Anything beyond that - a human-written comment on what a
helper IP actually is, or the subnet/gateway details of the network *that
IP* lives on - would need an external IPAM/DDI system (e.g. Infoblox) that
this app has no connection to, so those fields are intentionally not
fabricated here.
"""
from __future__ import annotations

import re
import socket
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=14)
SUBTITLE_FONT = Font(name=FONT, size=10, italic=True, color="555555")
BODY_FONT = Font(name=FONT, size=10)
BOLD_BODY = Font(name=FONT, size=10, bold=True)
MONO_FONT = Font(name="Consolas", size=10)
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fetch_running_config(host: str, port: int, device_type: str, username: str, password: str, enable_secret: str, timeout: int = 60) -> str:
    """Connects to a single device and returns 'show running-config' output.
    Deliberately minimal compared to collector.collect_one() - DDI only
    ever needs one command from one device synchronously, not the full
    multi-device job/progress/cancellation machinery. Importing collector
    also applies its ssh-rsa compatibility patch, so this benefits from
    that too."""
    from netmiko import ConnectHandler

    from . import collector  # noqa: F401  (applies the paramiko ssh-rsa patch on import)

    connection = ConnectHandler(
        device_type=device_type, host=host, port=port, username=username, password=password,
        secret=enable_secret, conn_timeout=timeout, auth_timeout=timeout, banner_timeout=timeout,
        fast_cli=False,
    )
    try:
        if not connection.check_enable_mode():
            connection.enable()
        return connection.send_command("show running-config", read_timeout=90)
    finally:
        connection.disconnect()


def parse_vlan_interfaces(config_text: str) -> list[dict[str, Any]]:
    """Parses every 'interface Vlan<N>' block out of a running-config dump.
    Each block runs from its 'interface VlanN' line to the next bare '!'
    line (standard IOS block-termination), so unrelated interfaces and
    global config are never absorbed. A VLAN with no configured IP or
    helper addresses (e.g. an unused/shutdown VLAN 1) still produces an
    entry, with those fields simply blank - matching how such VLANs still
    appear as their own row in the reference workbook this was modeled on.
    """
    results: list[dict[str, Any]] = []
    for match in re.finditer(r"(?m)^interface Vlan(\d+)\s*$", config_text):
        vlan_num = match.group(1)
        start = match.end()
        end_match = re.search(r"(?m)^!\s*$", config_text[start:])
        block = config_text[start:start + end_match.start()] if end_match else config_text[start:]
        description = re.search(r"(?m)^\s*description\s+(.+)$", block)
        ip_addr = re.search(r"(?m)^\s*ip address\s+(\S+)\s+(\S+)", block)
        helpers = re.findall(r"(?m)^\s*ip helper-address\s+(\S+)", block)
        results.append({
            "vlan": int(vlan_num), "interface": f"interface Vlan{vlan_num}",
            "description": description.group(1).strip() if description else "",
            "svi_ip": ip_addr.group(1) if ip_addr else "",
            "svi_mask": ip_addr.group(2) if ip_addr else "",
            "helpers": helpers,
            "raw_block": block.strip("\n"),
        })
    return results


def build_vlan_helpers_rows(vlan_interfaces: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Flattens to one row per (VLAN, helper) pair - a VLAN with several
    helpers gets several rows; a VLAN with none still gets exactly one row
    with a blank helper address, so every VLAN interface is represented."""
    rows: list[dict[str, Any]] = []
    for vlan in vlan_interfaces:
        base = {
            "vlan": vlan["vlan"], "interface": vlan["interface"],
            "description": vlan["description"], "svi_ip": vlan["svi_ip"],
        }
        if vlan["helpers"]:
            for helper in vlan["helpers"]:
                rows.append({**base, "helper_address": helper})
        else:
            rows.append({**base, "helper_address": ""})
    return rows


def resolve_helper_hostname(ip: str) -> str:
    """Best-effort reverse DNS lookup for a helper IP. Returns an empty
    string (not the IP itself) if it doesn't resolve, so it's visually
    obvious in the sheet which entries came from a real PTR record."""
    try:
        return socket.gethostbyaddr(ip)[0]
    except (socket.herror, socket.gaierror, OSError):
        return ""


def build_unique_helpers_rows(vlan_interfaces: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Deduplicates every helper IP seen across all VLANs, with how many
    VLANs reference it and a best-effort reverse-DNS hostname. Does not
    attempt a comment, gateway, or mask for the helper's own network -
    see the module docstring for why."""
    counts: dict[str, int] = {}
    order: list[str] = []
    for vlan in vlan_interfaces:
        for helper in vlan["helpers"]:
            if helper not in counts:
                order.append(helper)
            counts[helper] = counts.get(helper, 0) + 1
    return [
        {"helper_address": ip, "hostname": resolve_helper_hostname(ip), "vlan_count": counts[ip]}
        for ip in order
    ]


def build_svi_config_text(vlan_interfaces: list[dict[str, Any]]) -> str:
    """A condensed text view: for each VLAN, just the interface, description,
    and helper-address lines - not the full block (no ACLs, HSRP, etc.) -
    matching the 'SVI Condensed Config' sheet in the reference workbook."""
    lines: list[str] = []
    for vlan in vlan_interfaces:
        lines.append(vlan["interface"])
        if vlan["description"]:
            lines.append(f" description {vlan['description']}")
        for helper in vlan["helpers"]:
            lines.append(f" ip helper-address {helper}")
        lines.append("!")
    return "\n".join(lines)


def analyze_config(
    config_text: str,
    source_label: str,
    hide_no_helpers: bool = False,
    exclude_vlans: set[int] | None = None,
) -> dict[str, Any]:
    """hide_no_helpers drops VLAN interfaces with zero configured helper
    addresses - useful once you're specifically after DHCP relay info and
    don't care about VLANs that aren't relaying at all. exclude_vlans drops
    specific VLAN numbers outright regardless of whether they have
    helpers - e.g. a Guest VLAN whose DHCP server is intentionally out of
    scope for a given project. Both are general filters, not tied to any
    particular VLAN number or site; every downstream view (the flattened
    table, the unique-helper counts, the condensed config text) is built
    from the already-filtered list, so a filtered-out VLAN's helpers don't
    still show up counted in Unique Helpers."""
    exclude_vlans = exclude_vlans or set()
    vlan_interfaces = parse_vlan_interfaces(config_text)
    vlan_interfaces = [
        v for v in vlan_interfaces
        if v["vlan"] not in exclude_vlans and not (hide_no_helpers and not v["helpers"])
    ]
    return {
        "source_label": source_label,
        "vlan_interfaces": vlan_interfaces,
        "vlan_helpers": build_vlan_helpers_rows(vlan_interfaces),
        "unique_helpers": build_unique_helpers_rows(vlan_interfaces),
        "svi_config_text": build_svi_config_text(vlan_interfaces),
    }


def _write_sheet(ws, headers: list[str], rows: list[list[Any]], widths: list[int]) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for r, row in enumerate(rows, start=2):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = BODY_FONT
            cell.border = BORDER
    if rows:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"


def build_ddi_workbook(analysis: dict[str, Any]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 95
    ws["A1"] = "DDI Analysis - VLAN / DHCP Helper Addresses"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Source: {analysis['source_label']}  |  Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = SUBTITLE_FONT
    lines = [
        "", "WHAT THIS COVERS",
        "VLAN Helpers, SVI Config, and the helper-address list/counts below are all parsed directly",
        "from the device's own configuration. Each helper IP's hostname (where shown) is a live",
        "reverse-DNS lookup, not a config value - it will be blank if that IP has no PTR record",
        "reachable from wherever this app runs.",
        "",
        "NOT INCLUDED",
        "A human-readable comment on what each helper IP actually is, and that IP's own",
        "gateway/subnet mask, would need an external IPAM/DDI system (e.g. Infoblox) - this app has",
        "no connection to one, so those fields are intentionally left out rather than guessed at.",
        "",
        f"VLAN interfaces found: {len(analysis['vlan_interfaces'])}",
        f"Unique helper addresses: {len(analysis['unique_helpers'])}",
    ]
    for i, line in enumerate(lines, start=3):
        cell = ws.cell(row=i, column=1, value=line)
        cell.font = BOLD_BODY if line.isupper() else BODY_FONT

    vlan_helpers_rows = [
        [r["vlan"], r["interface"], r["description"], r["svi_ip"], r["helper_address"]]
        for r in analysis["vlan_helpers"]
    ]
    ws2 = wb.create_sheet("VLAN Helpers")
    _write_sheet(ws2, ["VLAN", "Interface", "Description", "SVI IP", "Helper Address"],
                 vlan_helpers_rows, [10, 20, 26, 18, 20])

    unique_rows = [
        [r["helper_address"], r["hostname"], r["vlan_count"]]
        for r in analysis["unique_helpers"]
    ]
    ws3 = wb.create_sheet("Unique Helpers")
    _write_sheet(ws3, ["Helper Address", "Hostname (reverse DNS)", "VLANs Using It"],
                 unique_rows, [20, 34, 16])

    ws4 = wb.create_sheet("SVI Config")
    ws4.column_dimensions["A"].width = 90
    for i, line in enumerate(analysis["svi_config_text"].splitlines(), start=1):
        cell = ws4.cell(row=i, column=1, value=line)
        cell.font = MONO_FONT

    return wb