"""Compliance checking against customer standards.

Phase 1 (this module, for now): firmware compliance. Given a completed
collection job's devices (each with a model and a collected software
version) and the customer's Tech Stack workbook (a model-keyed table of
certified firmware versions), flags each device as compliant, non-compliant,
or unknown (model not found in the reference table).

Deliberately scoped to firmware only for now - SNMP/syslog compliance
against the 32-sheet Standard Configuration Templates workbook is a
follow-up that needs a resolved Model -> Role -> template-sheet mapping
first (several roles there don't map 1:1 to a sheet name), and ACL/QoS
compliance is a further step beyond that given how structurally complex
those sections are (policy-maps referencing class-maps referencing ACLs).
"""
from __future__ import annotations

import io
import re
from typing import Any

from openpyxl import load_workbook


def normalize_version(value: str) -> str:
    """Lowercases, strips whitespace, and removes leading zeros within each
    numeric run so '17.09.05' and '17.9.5' - both seen in the same Tech
    Stack cell as alternate notations of the same certified version -
    compare as equal, without treating genuinely different versions
    (17.9.4a vs 17.9.5) as a match."""
    value = value.strip().lower()
    return re.sub(r"\d+", lambda m: str(int(m.group())), value)


def parse_tech_stack(file_bytes: bytes) -> dict[str, dict[str, Any]]:
    """Parses the 'Tech Stack' sheet into a dict keyed by uppercased model
    number. Each entry holds the raw certified-version cell (which may
    contain several newline-separated acceptable notations of the same
    version), the role, and device type. If the same model appears more
    than once with genuinely different certified versions (rare - only
    seen for a couple of AP models where CAPWAP vs Autonomous mode expects
    different firmware), all of the distinct versions are kept and the
    entry is marked ambiguous, rather than silently picking one."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    if "Tech Stack" not in wb.sheetnames:
        return {}
    ws = wb["Tech Stack"]
    rows = list(ws.iter_rows(values_only=True))

    header_idx = None
    for i, row in enumerate(rows):
        if row and row[0] == "Device Type" and row[2] == "Model":
            header_idx = i
            break
    if header_idx is None:
        return {}

    result: dict[str, dict[str, Any]] = {}
    for row in rows[header_idx + 1:]:
        if not row or not row[2]:
            continue
        model = str(row[2]).strip().upper()
        certified_raw = row[6] if len(row) > 6 else None
        if not certified_raw:
            continue
        notations = [n.strip() for n in str(certified_raw).splitlines() if n.strip()]
        entry = result.setdefault(model, {
            "device_type": row[0], "role": row[3], "notations": set(), "ambiguous": False,
        })
        existing_normalized = {normalize_version(n) for n in entry["notations"]}
        new_normalized = {normalize_version(n) for n in notations}
        if existing_normalized and not (existing_normalized & new_normalized):
            entry["ambiguous"] = True  # a real conflict, not just a second identical listing
        entry["notations"].update(notations)
    return result


def check_firmware_compliance(devices: list[dict[str, Any]], tech_stack: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    """devices: the collection job's own result records (each needs
    inventory_name/hostname, model, software_version - exactly what's
    already in a job's `results`). Returns one row per device with a
    compliance verdict; devices with no model recorded, or a model not
    found in the Tech Stack, are marked as unknown rather than silently
    skipped, so nothing goes unaccounted for in the report."""
    rows: list[dict[str, Any]] = []
    for device in devices:
        model = (device.get("model") or "").strip().upper()
        actual_version = (device.get("software_version") or "").strip()
        hostname = device.get("detected_hostname") or device.get("inventory_name") or ""
        base = {
            "hostname": hostname, "host": device.get("host", ""), "model": device.get("model") or "",
            "role": device.get("role") or "", "actual_version": actual_version,
        }
        if not model:
            rows.append({**base, "status": "UNKNOWN", "expected_version": "",
                         "detail": "No model recorded for this device (collection may have failed)."})
            continue
        entry = tech_stack.get(model)
        if entry is None:
            rows.append({**base, "status": "UNKNOWN", "expected_version": "",
                         "detail": f"Model '{device.get('model')}' not found in the Tech Stack reference."})
            continue
        expected_display = " or ".join(sorted(entry["notations"]))
        if not actual_version:
            rows.append({**base, "status": "UNKNOWN", "expected_version": expected_display,
                         "detail": "No software version was collected for this device."})
            continue
        actual_normalized = normalize_version(actual_version)
        is_match = any(normalize_version(n) == actual_normalized for n in entry["notations"])
        detail = ""
        if entry["ambiguous"]:
            detail = "Note: this model has more than one certified version on record depending on role/mode - verify manually."
        rows.append({
            **base, "expected_version": expected_display,
            "status": "COMPLIANT" if is_match else "NON_COMPLIANT",
            "detail": detail,
        })
    return rows
