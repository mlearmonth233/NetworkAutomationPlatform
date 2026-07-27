"""R2O Check: cross-references the 5 standard R2O deliverables (LLD, Network
Diagram, Rack Elevations, CMDB export, IT/OT Sitebook) against the Sitebook
as source of truth, flagging hostname/MAC/serial/asset-tag/model/IP
inconsistencies.

Built and validated against one real site's complete document set. The row
and column detection below uses structural signals - sheet names, column
header text, the '!' row marker in the LLD's Equipment sheet - rather than
anything specific to that site's hostname naming, since those signals looked
like they come from a standardized template rather than a one-off layout.
That said, spreadsheet templates do vary: if a different site's documents
use a meaningfully different sheet/column layout, a source may contribute
fewer hostnames than expected, or none at all. A source reporting 0
hostnames in the result is a signal to check that document's structure
against what this module expects, not necessarily a real finding.
"""
from __future__ import annotations

import io
import re
import xml.etree.ElementTree as ET
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

FIELDS = ("mac", "serial", "asset_tag", "model", "ip")

# Which fields each source's own schema actually carries - a field absent
# here is never flagged as "missing" for that source, since the document
# type was never designed to hold it (e.g. a network diagram has no MAC
# address column at all).
FIELDS_TRACKED = {
    "LLD-Equipment": {"model", "ip"},
    "LLD-Wireless": {"mac", "serial", "asset_tag", "model", "ip"},
    "CMDB": {"mac", "serial", "asset_tag", "model", "ip"},
    "NetworkDiagram": {"model", "ip"},
    "RackElevations": {"model", "ip"},
}

SOURCE_LABELS = {
    "LLD-Equipment": "LLD (Equipment)",
    "LLD-Wireless": "LLD (Wireless)",
    "CMDB": "CMDB Export",
    "NetworkDiagram": "Network Diagram",
    "RackElevations": "Rack Elevations",
}


# ---------------------------------------------------------------------------
# Normalization helpers
# ---------------------------------------------------------------------------

def norm_mac(value: Any) -> str | None:
    if not value:
        return None
    hexonly = re.sub(r"[^0-9a-fA-F]", "", str(value))
    return hexonly.lower() if len(hexonly) == 12 else None


def norm_text(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def norm_for_compare(field_name: str, value: str) -> str:
    """Model numbers get compared case/punctuation-insensitively so cosmetic
    differences ('C9300-24UX' vs 'c9300 24ux') aren't reported as conflicts."""
    if field_name == "model" and value:
        return re.sub(r"[\s\-_]+", "", value).lower()
    return value


def _is_plausible_hostname(value: str | None) -> bool:
    """Deliberately generic - no site-specific prefix assumptions. Excludes
    the most common ways a non-device row leaks into a hostname column:
    footnote/legend text (spaces, leading punctuation) and obviously wrong
    lengths."""
    if not value:
        return False
    if " " in value or value.startswith(("*", "!", "#")):
        return False
    if len(value) < 3 or len(value) > 64:
        return False
    return True


def _sheet_rows(wb, sheet_name: str):
    if sheet_name not in wb.sheetnames:
        return None, []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None, []
    return rows[0], rows[1:]


def _col_index(header, *names) -> int | None:
    for name in names:
        if header and name in header:
            return header.index(name)
    return None


# ---------------------------------------------------------------------------
# Per-document parsers - each takes raw file bytes, returns hostname-keyed dicts
# ---------------------------------------------------------------------------

def parse_sitebook(data: bytes) -> dict[str, dict]:
    """IT/OT Sitebook. Reads 'Devices and Serials' (wired gear), 'AP Info'
    (wireless APs), and 'PDU Info' (power distribution) - merged into one
    hostname-keyed dict. A missing sheet is skipped, not fatal."""
    wb = load_workbook(io.BytesIO(data), read_only=True, keep_vba=True, data_only=True)
    out: dict[str, dict] = {}

    header, rows = _sheet_rows(wb, "Devices and Serials")
    if header:
        idx = {name: i for i, name in enumerate(header) if name}
        host_i = _col_index(header, "Host", "Hostname")
        for row in rows:
            host = norm_text(row[host_i]) if host_i is not None else None
            if not _is_plausible_hostname(host):
                continue
            out[host] = {
                "mac": norm_mac(row[idx["Mac Address"]]) if "Mac Address" in idx else None,
                "serial": norm_text(row[idx["SN"]]) if "SN" in idx else None,
                "asset_tag": norm_text(row[idx["Asset Tag"]]) if "Asset Tag" in idx else None,
                "model": norm_text(row[idx["PID"]]) if "PID" in idx else None,
                "ip": norm_text(row[idx["IP"]]) if "IP" in idx else None,
            }

    header, rows = _sheet_rows(wb, "AP Info")
    if header:
        idx = {name: i for i, name in enumerate(header) if name}
        host_i = _col_index(header, "Cisco AP Name", "AP Name", "Hostname")
        for row in rows:
            host = norm_text(row[host_i]) if host_i is not None else None
            if not _is_plausible_hostname(host):
                continue
            out[host] = {
                "mac": norm_mac(row[idx["MAC Address"]]) if "MAC Address" in idx else None,
                "serial": norm_text(row[idx["Serial Number"]]) if "Serial Number" in idx else None,
                "asset_tag": norm_text(row[idx["Asset Tag"]]) if "Asset Tag" in idx else None,
                "model": norm_text(row[idx["AP Model"]]) if "AP Model" in idx else None,
                "ip": norm_text(row[idx["IP Address"]]) if "IP Address" in idx else None,
            }

    header, rows = _sheet_rows(wb, "PDU Info")
    if header:
        idx = {name: i for i, name in enumerate(header) if name}
        host_i = _col_index(header, "Host", "Hostname")
        for row in rows:
            host = norm_text(row[host_i]) if host_i is not None else None
            if not _is_plausible_hostname(host):
                continue
            out[host] = {
                "mac": norm_mac(row[idx["MAC Address"]]) if "MAC Address" in idx else None,
                "serial": norm_text(row[idx["Serial Number"]]) if "Serial Number" in idx else None,
                "asset_tag": norm_text(row[idx["Asset Tag"]]) if "Asset Tag" in idx else None,
                "model": norm_text(row[idx["Model Number"]]) if "Model Number" in idx else None,
                "ip": norm_text(row[idx["IP Address"]]) if "IP Address" in idx else None,
            }
    return out


def parse_lld(data: bytes) -> dict[str, dict[str, dict]]:
    """LLD. 'Equipment' (wired gear - hostname/platform/ip only, no
    mac/serial/asset tag by design) and 'Wireless' (AP mac/serial/asset
    tag/model/ip). Returns two separately-labeled sources so each is
    compared against the Sitebook independently."""
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    result: dict[str, dict] = {"LLD-Equipment": {}, "LLD-Wireless": {}}

    header, rows = _sheet_rows(wb, "Equipment")
    if header:
        for row in rows:
            marker = row[0] if len(row) > 0 else None
            host = norm_text(row[1]) if len(row) > 1 else None
            model = norm_text(row[2]) if len(row) > 2 else None
            ip = norm_text(row[6]) if len(row) > 6 else None
            # Real device rows are marked '!' in column A. This sheet also
            # contains section-header rows (e.g. 'DC-1', 'IDF-3') that share
            # the same '!' marker but carry no model/IP data at all, plus a
            # second, differently-shaped summary/tally table further down
            # (no '!' marker) - requiring at least one populated data field
            # excludes both.
            if marker != "!" or not _is_plausible_hostname(host) or host in ("Hostname", "Device"):
                continue
            if not model and not ip:
                continue
            result["LLD-Equipment"][host] = {
                "mac": None, "serial": None, "asset_tag": None,
                "model": model,
                "ip": ip,
            }

    header, rows = _sheet_rows(wb, "Wireless")
    if header:
        for row in rows:
            host = norm_text(row[1]) if len(row) > 1 else None
            mac_raw = row[2] if len(row) > 2 else None
            if not _is_plausible_hostname(host) or host == "Hostname" or mac_raw == "!":
                continue
            result["LLD-Wireless"][host] = {
                "mac": norm_mac(row[2]) if len(row) > 2 else None,
                "serial": norm_text(row[3]) if len(row) > 3 else None,
                "asset_tag": norm_text(row[4]) if len(row) > 4 else None,
                "model": norm_text(row[5]) if len(row) > 5 else None,
                "ip": norm_text(row[6]) if len(row) > 6 else None,
            }
    return result


def parse_cmdb(data: bytes) -> tuple[dict, set]:
    """CMDB export's 'CMDB Update' sheet. Returns (active_records,
    decommissioned_hostnames) - 'Decomm' rows are expected to be absent
    from the current Sitebook (they're being retired), so they're tracked
    separately rather than flagged as a coverage gap."""
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    active: dict[str, dict] = {}
    decommissioned: set[str] = set()

    header, rows = _sheet_rows(wb, "CMDB Update")
    if not header:
        return active, decommissioned
    idx = {name: i for i, name in enumerate(header) if name}
    if "Name" not in idx or "Action" not in idx:
        return active, decommissioned

    for row in rows:
        host = norm_text(row[idx["Name"]])
        action = norm_text(row[idx["Action"]])
        if not _is_plausible_hostname(host):
            continue
        if action == "Decomm":
            decommissioned.add(host)
            continue
        if action not in ("Add", "Add Spare"):
            continue
        active[host] = {
            "mac": norm_mac(row[idx["MAC Address"]]) if "MAC Address" in idx else None,
            "serial": norm_text(row[idx["Serial Number"]]) if "Serial Number" in idx else None,
            "asset_tag": norm_text(row[idx["Asset Tag"]]) if "Asset Tag" in idx else None,
            "model": norm_text(row[idx["Model ID"]]) if "Model ID" in idx else None,
            "ip": norm_text(row[idx["IP Address"]]) if "IP Address" in idx else None,
        }
    return active, decommissioned


def parse_visio_linked_data(data: bytes) -> dict[str, dict]:
    """Extracts hostname/IP/model from a .vsdx's linked data tables (Visio's
    'Link Data to Shapes' feature) if present - a diagram built this way
    carries a cached copy of whatever spreadsheet was linked in, which is
    far more reliable to read than trying to parse visual shape text.
    Returns {} for a diagram built without linked data."""
    out: dict[str, dict] = {}
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            recordset_names = [n for n in zf.namelist() if re.match(r"visio/data/recordset\d+\.xml$", n)]
            for name in recordset_names:
                try:
                    root = ET.fromstring(zf.read(name))
                except ET.ParseError:
                    continue
                # Not every column keeps a clean attribute name in the row
                # elements themselves (some become c0, c1, ... with the
                # real label only present in the schema's rs:name) - build
                # that mapping from the schema first.
                label_by_attr: dict[str, str] = {}
                for attr in root.iter("{uuid:BDC6E3F0-6DA3-11d1-A2A3-00AA00C14882}AttributeType"):
                    attr_name = attr.get("name")
                    label = attr.get("{urn:schemas-microsoft-com:rowset}name") or attr_name
                    if attr_name:
                        label_by_attr[attr_name] = label
                for row in root.findall(".//{#RowsetSchema}row"):
                    values = {label_by_attr.get(k, k): v for k, v in row.attrib.items()}
                    host = norm_text(values.get("Hostname"))
                    if not _is_plausible_hostname(host):
                        continue
                    out[host] = {
                        "mac": None, "serial": None, "asset_tag": None,
                        "model": norm_text(values.get("Unit 1") or values.get("PDU1")),
                        "ip": norm_text(values.get("IP")),
                    }
    except zipfile.BadZipFile:
        pass
    return out


# ---------------------------------------------------------------------------
# Cross-reference
# ---------------------------------------------------------------------------

def cross_reference(sitebook: dict, others: dict[str, dict], decommissioned: set) -> dict[str, list]:
    conflicts: list[dict] = []
    sitebook_gaps: list[dict] = []
    coverage_gaps: list[dict] = []

    all_hostnames = set(sitebook)
    for data in others.values():
        all_hostnames |= set(data)

    for host in sorted(all_hostnames):
        sb = sitebook.get(host)
        if sb is None:
            if host in decommissioned:
                continue
            present_in = [src for src, data in others.items() if host in data]
            if present_in:
                coverage_gaps.append({"hostname": host, "present_in": present_in})
            continue

        for src, data in others.items():
            other = data.get(host)
            if other is None:
                continue
            tracked = FIELDS_TRACKED.get(src, set(FIELDS))
            for f in FIELDS:
                if f not in tracked:
                    continue
                sb_val = sb.get(f)
                other_val = other.get(f)
                if sb_val and other_val:
                    if norm_for_compare(f, sb_val) != norm_for_compare(f, other_val):
                        conflicts.append({
                            "hostname": host, "field": f, "source": src,
                            "sitebook_value": sb_val, "other_value": other_val,
                        })
                elif (not sb_val) and other_val:
                    sitebook_gaps.append({
                        "hostname": host, "field": f, "source": src, "other_value": other_val,
                    })

    return {"conflicts": conflicts, "sitebook_gaps": sitebook_gaps, "coverage_gaps": coverage_gaps}


def _dominant_prefix(hostnames, prefix_len: int = 8) -> str | None:
    if not hostnames:
        return None
    counts: dict[str, int] = {}
    for h in hostnames:
        counts[h[:prefix_len]] = counts.get(h[:prefix_len], 0) + 1
    return max(counts.items(), key=lambda kv: kv[1])[0]


def _group_coverage_gaps(coverage_gaps: list[dict], sitebook_hostnames, min_group_size: int = 5) -> list[dict]:
    """Groups coverage gaps by hostname prefix (first 8 characters, matching
    a common site+subcode length) to surface a systematic block of missing
    devices as one finding. Only a prefix that DIFFERS from the Sitebook's
    own dominant prefix is flagged this way - several unrelated gaps that
    happen to share the site's normal naming convention aren't a distinct
    block, just individually scattered gaps."""
    dominant = _dominant_prefix(sitebook_hostnames)
    by_prefix: dict[str, list[str]] = {}
    for g in coverage_gaps:
        prefix = g["hostname"][:8]
        by_prefix.setdefault(prefix, []).append(g["hostname"])
    return [
        {"prefix": prefix, "hostnames": hosts, "count": len(hosts)}
        for prefix, hosts in sorted(by_prefix.items(), key=lambda kv: -len(kv[1]))
        if len(hosts) >= min_group_size and prefix != dominant
    ]


def _structural_gaps(sitebook: dict, others: dict[str, dict], sitebook_gaps: list[dict]) -> list[dict]:
    """Detects when a (source, field) gap affects nearly every hostname that
    source has in common with the Sitebook - a sign the Sitebook simply
    doesn't capture that field for that device category at all, rather than
    a per-device data-quality problem worth listing individually."""
    counts: dict[tuple[str, str], int] = {}
    for g in sitebook_gaps:
        counts[(g["source"], g["field"])] = counts.get((g["source"], g["field"]), 0) + 1

    findings = []
    for (src, f), gap_count in counts.items():
        overlap = len(set(sitebook) & set(others.get(src, {})))
        if overlap and gap_count / overlap >= 0.9 and gap_count >= 5:
            findings.append({
                "source": src, "field": f, "gap_count": gap_count, "overlap": overlap,
            })
    return findings


# ---------------------------------------------------------------------------
# Report generation
# ---------------------------------------------------------------------------

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=14)
SUBTITLE_FONT = Font(name=FONT, size=10, italic=True, color="555555")
BODY_FONT = Font(name=FONT, size=10)
BOLD_BODY = Font(name=FONT, size=10, bold=True)
RED_FILL = PatternFill("solid", fgColor="FDE2E1")
AMBER_FILL = PatternFill("solid", fgColor="FEF3C7")
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build_summary(
    findings: dict[str, list],
    sitebook: dict,
    others: dict[str, dict],
    decommissioned: set,
) -> dict[str, Any]:
    """Computes the same grouped/structural summary used in the report's
    Summary sheet, as structured data - so an API response can render an
    inline summary without needing to parse it back out of the workbook."""
    grouped_gaps = _group_coverage_gaps(findings["coverage_gaps"], sitebook.keys())
    structural = _structural_gaps(sitebook, others, findings["sitebook_gaps"])
    return {
        "conflict_count": len(findings["conflicts"]),
        "coverage_gap_count": len(findings["coverage_gaps"]),
        "decommissioned_count": len(decommissioned),
        "systematic_groups": grouped_gaps,
        "structural_gaps": [
            {
                "source": SOURCE_LABELS.get(s["source"], s["source"]),
                "field": s["field"],
                "gap_count": s["gap_count"],
                "overlap": s["overlap"],
            }
            for s in structural
        ],
    }


def build_report_workbook(
    findings: dict[str, list],
    sitebook: dict,
    others: dict[str, dict],
    decommissioned: set,
    site_label: str,
    sources_provided: list[str],
    sources_missing: list[str],
    summary: dict[str, Any],
) -> Workbook:
    conflicts = findings["conflicts"]
    coverage_gaps = findings["coverage_gaps"]
    grouped_gaps = summary["systematic_groups"]
    grouped_hostnames = {h for g in grouped_gaps for h in g["hostnames"]}
    structural = summary["structural_gaps"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 95
    ws["A1"] = f"R2O Data Consistency Check - {site_label}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Source of truth: IT/OT Sitebook"
    ws["A2"].font = SUBTITLE_FONT

    lines: list[tuple[str, Font]] = [
        ("", BODY_FONT),
        ("DOCUMENTS PROVIDED", BOLD_BODY),
        (f"  {', '.join(sources_provided) if sources_provided else '(none)'}", BODY_FONT),
    ]
    if sources_missing:
        lines.append((f"  Not provided (skipped): {', '.join(sources_missing)}", BODY_FONT))
    lines += [
        ("", BODY_FONT),
        ("RESULTS AT A GLANCE", BOLD_BODY),
        (f"  {len(conflicts)} value conflicts (same device, genuinely different value between documents)", BODY_FONT),
        (f"  {len(coverage_gaps)} devices present in another document but missing from the Sitebook entirely", BODY_FONT),
        (f"  {len(decommissioned)} CMDB 'Decomm' entries excluded from the gap check (expected to be absent)", BODY_FONT),
    ]

    if grouped_gaps:
        lines.append(("", BODY_FONT))
        lines.append(("SYSTEMATIC GAPS - LIKELY WORTH CHECKING FIRST", BOLD_BODY))
        for g in grouped_gaps:
            lines.append((
                f"  {g['count']} devices starting with '{g['prefix']}' exist in other documents "
                f"but do not appear in the Sitebook at all.", BODY_FONT,
            ))

    if structural:
        lines.append(("", BODY_FONT))
        lines.append(("STRUCTURAL GAPS (a field the Sitebook doesn't capture for this device category - not per-device errors)", BOLD_BODY))
        for s in structural:
            lines.append((
                f"  {s['source']}: '{s['field']}' is populated for "
                f"{s['gap_count']} of {s['overlap']} matching devices, but the Sitebook has none of them.",
                BODY_FONT,
            ))

    if conflicts:
        lines.append(("", BODY_FONT))
        lines.append(("SEE THE 'CONFLICTS' TAB FOR THE FULL LIST OF VALUE MISMATCHES", BOLD_BODY))

    row = 3
    for text, font in lines:
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = font
        row += 1

    # Conflicts sheet
    ws2 = wb.create_sheet("Conflicts")
    headers = ["Hostname", "Field", "Sitebook Value", "Conflicting Source", "Conflicting Value"]
    for col, h in enumerate(headers, start=1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
    for col, w in enumerate([20, 12, 26, 20, 26], start=1):
        ws2.column_dimensions[get_column_letter(col)].width = w
    r = 2
    for c in sorted(conflicts, key=lambda x: (x["hostname"], x["field"])):
        values = [c["hostname"], c["field"], c["sitebook_value"], SOURCE_LABELS.get(c["source"], c["source"]), c["other_value"]]
        for col, v in enumerate(values, start=1):
            cell = ws2.cell(row=r, column=col, value=v)
            cell.font = BODY_FONT
            cell.border = BORDER
        r += 1
    if r > 2:
        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = f"A1:E{r - 1}"

    # Coverage Gaps sheet
    ws3 = wb.create_sheet("Coverage Gaps")
    headers = ["Hostname", "Present In (not in Sitebook)", "Part of a systematic group?"]
    for col, h in enumerate(headers, start=1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
    ws3.column_dimensions["A"].width = 24
    ws3.column_dimensions["B"].width = 45
    ws3.column_dimensions["C"].width = 40
    r = 2
    for g in sorted(coverage_gaps, key=lambda x: x["hostname"]):
        in_group = g["hostname"] in grouped_hostnames
        note = f"Yes - see Summary ({g['hostname'][:8]} group)" if in_group else ""
        values = [g["hostname"], ", ".join(SOURCE_LABELS.get(s, s) for s in g["present_in"]), note]
        for col, v in enumerate(values, start=1):
            cell = ws3.cell(row=r, column=col, value=v)
            cell.font = BODY_FONT
            cell.border = BORDER
            if in_group:
                cell.fill = RED_FILL
        r += 1
    if r > 2:
        ws3.freeze_panes = "A2"
        ws3.auto_filter.ref = f"A1:C{r - 1}"

    return wb


def run_r2o_check(
    site_label: str,
    sitebook_bytes: bytes | None,
    lld_bytes: bytes | None,
    cmdb_bytes: bytes | None,
    network_diagram_bytes: bytes | None,
    rack_elevations_bytes: bytes | None,
) -> tuple[dict[str, list], dict[str, Any], Workbook, dict[str, int]]:
    """Runs the full pipeline given whichever documents were provided (all
    are optional - anything not supplied is simply skipped, and noted in the
    report). Returns (findings, summary, workbook, source_counts)."""
    sources_provided = []
    sources_missing = []

    sitebook = parse_sitebook(sitebook_bytes) if sitebook_bytes else {}
    (sources_provided if sitebook_bytes else sources_missing).append("IT/OT Sitebook")

    others: dict[str, dict] = {}
    decommissioned: set[str] = set()

    if lld_bytes:
        lld_result = parse_lld(lld_bytes)
        others.update(lld_result)
        sources_provided.append("LLD")
    else:
        sources_missing.append("LLD")

    if cmdb_bytes:
        cmdb_active, decommissioned = parse_cmdb(cmdb_bytes)
        others["CMDB"] = cmdb_active
        sources_provided.append("CMDB Export")
    else:
        sources_missing.append("CMDB Export")

    if network_diagram_bytes:
        others["NetworkDiagram"] = parse_visio_linked_data(network_diagram_bytes)
        sources_provided.append("Network Diagram")
    else:
        sources_missing.append("Network Diagram")

    if rack_elevations_bytes:
        others["RackElevations"] = parse_visio_linked_data(rack_elevations_bytes)
        sources_provided.append("Rack Elevations")
    else:
        sources_missing.append("Rack Elevations")

    findings = cross_reference(sitebook, others, decommissioned)
    summary = build_summary(findings, sitebook, others, decommissioned)
    workbook = build_report_workbook(
        findings, sitebook, others, decommissioned, site_label, sources_provided, sources_missing, summary,
    )
    source_counts = {"Sitebook": len(sitebook), **{k: len(v) for k, v in others.items()}}
    return findings, summary, workbook, source_counts
